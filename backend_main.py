"""
Spinnaker Rental Model — Backend API
Excel file is the single source of truth. This layer only reads/writes green cells
and recalculates via LibreOffice. No formulas are reimplemented here.
"""
import shutil, subprocess, tempfile, os, json
from pathlib import Path
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import load_workbook

app = FastAPI(title="Spinnaker Rental Model API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ─── SINGLE SOURCE OF TRUTH: Excel file path ───────────────────────────────────
EXCEL_TEMPLATE = Path(__file__).parent / "Spinnakeri_uürimudel_v3.xlsx"
SHEET = "Spinnakeri mudel"

# ─── INPUT CELL MAP  (UI field → Excel cell address) ──────────────────────────
# Only green cells (theme fill 9) are ever written. Nothing else is touched.
INPUT_CELLS = {
    # Financing parameters — Base column (C)
    "equity_share":           "C5",   # Omakapitali osakaal
    "loan_interest":          "C6",   # Laenuintress
    "repayment_years":        "C7",   # Tagasimakse graafik annuiteedina (a)
    "project_duration":       "C8",   # Projekti kestvus (a)

    # Property parameters — Target column (D)
    "vat_rate":               "D12",  # Käibemaks
    "num_apartments":         "D13",  # Korterite arv
    "avg_apt_size_m2":        "D14",  # Keskmine korteri pind m2
    "land_price_per_m2":      "D19",  # Maa hind (kogu kinnistu m2 hind)
    "planning_design_m2":     "D21",  # Planeering ja projekteerimine €/m²
    "connections_infra_m2":   "D22",  # Liitumistasud ja sotsiaalinfra €/m²
    "pm_supervision_m2":      "D23",  # Ehitusaegne projektijuhtimine €/m²
    "sales_service_m2":       "D24",  # Müük/klienditeenindus €/m²

    # Build costs — Target column (D) + ratios (E)
    "apts_build_cost_m2":     "D27",  # Korterid (üüritav osa) €/m²
    "apts_area_ratio":        "E27",  # Korterite osakaal pinnast
    "common_build_cost_m2":   "D28",  # Üldkasutatavad pinnad €/m²
    "common_area_ratio":      "E28",  # Üldkasutatavate pindade osakaal
    "parking_build_cost_m2":  "D29",  # Parkimiskorrus €/m²

    # Rent parameters
    "vacancy_credit_risk":    "D42",  # Vakants + makserisk osakaal NOIst
    "vacancy_first_year":     "E42",  # Esimene aasta vakants
    "opex_first_year_m2":     "D43",  # Halduskulu esimesel kasutusaastal m2 kohta
    "rent_increase_annual":   "D44",  # Üürihinna tõus aastas
    "monthly_rent_m2":        "D45",  # Esimese aasta kuuüür m2 kohta ← KEY INPUT

    # EIS measure toggle
    "eis_co_loan":            "D48",  # EISi kaaslaen: "Jah" or "Ei"
}

# ─── OUTPUT CELL MAP  (Excel cell → UI display field) ─────────────────────────
# These cells are ONLY read, never written.
OUTPUT_CELLS = {
    # IRR and equity
    "project_irr":            "H23",  # Projekti IRR
    "equity_irr":             "H24",  # Omakapitali IRR
    "max_equity_m2":          "H22",  # Omakapitali vajadus max m2 kohta
    "max_equity_eur":         "H26",  # Omakapitali vajadus max (eur)

    # Cost breakdown (total €)
    "development_cost":       "H28",  # ARENDUSKULU
    "vat_amount":             "H29",  # KÄIBEMAKS
    "opex_total":             "H30",  # HALDUSKULU
    "bank_return":            "H31",  # PANGA TOOTLUS
    "owner_return":           "H32",  # OMANIKU TOOTLUS
    "total_costs":            "H33",  # Kulud kokku
    "rental_income_total":    "H34",  # ÜÜRITULU
    "building_sale":          "H35",  # HOONE MÜÜK
    "total_revenues":         "H36",  # Tulud kokku

    # Area metrics
    "rentable_area_m2":       "D15",  # Üüritav pind kokku m2
    "total_net_area_m2":      "D16",  # Kogu suletud netopind m2

    # Cost per m² metrics
    "dev_cost_per_m2":        "D31",  # Arenduskulud kokku netopinna kohta
    "build_cost_per_m2":      "D32",  # Ehituskulud kokku netopinna kohta
    "total_cost_per_m2":      "D33",  # Arendus- ja ehituskulud kokku netopinna kohta
    "total_cost_incl_vat_m2": "D38",  # Total incl VAT per rentable m²
    "exit_value_m2":          "D39",  # Exit väärtus korteri m2 kohta

    # Effective financing params (post-EIS if active)
    "effective_equity_share": "D5",   # =IF(EIS="Jah", 10%, C5)
    "effective_interest":     "D6",   # =IF(EIS="Jah", C6-0.01, C6)
    "effective_repayment":    "D7",   # =IF(EIS="Jah", C7+5, C7)
}

# ─── Pydantic model for input payload ─────────────────────────────────────────
class ModelInputs(BaseModel):
    equity_share: float = 0.5
    loan_interest: float = 0.05
    repayment_years: int = 30
    project_duration: int = 30
    vat_rate: float = 0.24
    num_apartments: int = 40
    avg_apt_size_m2: float = 51.5
    land_price_per_m2: float = 20.0
    planning_design_m2: float = 80.0
    connections_infra_m2: float = 10.0
    pm_supervision_m2: float = 20.0
    sales_service_m2: float = 0.0
    apts_build_cost_m2: float = 1700.0
    apts_area_ratio: float = 0.8
    common_build_cost_m2: float = 1300.0
    common_area_ratio: float = 0.2
    parking_build_cost_m2: float = 1000.0
    vacancy_credit_risk: float = 0.05
    vacancy_first_year: float = 0.5
    opex_first_year_m2: float = 50.0
    rent_increase_annual: float = 0.02
    monthly_rent_m2: float = 15.0  # The primary pricing lever
    eis_co_loan: str = "Ei"  # "Jah" or "Ei"


def recalculate_excel(tmp_path: str) -> dict:
    """Trigger LibreOffice recalculation and return output values."""
    script = Path(__file__).parent.parent / "scripts" / "recalc.py"
    result = subprocess.run(
        ["python3", str(script), tmp_path, "30"],
        capture_output=True, text=True, timeout=60
    )
    recalc_result = json.loads(result.stdout)
    if recalc_result.get("status") == "errors_found":
        raise HTTPException(500, f"Excel recalc errors: {recalc_result['error_summary']}")

    wb = load_workbook(tmp_path, data_only=True)
    ws = wb[SHEET]
    outputs = {}
    for field, cell in OUTPUT_CELLS.items():
        val = ws[cell].value
        outputs[field] = float(val) if val is not None else None
    wb.close()
    return outputs


@app.get("/defaults")
def get_defaults():
    """Return current values of all green input cells from the template."""
    wb = load_workbook(EXCEL_TEMPLATE, data_only=True)
    ws = wb[SHEET]
    defaults = {}
    for field, cell in INPUT_CELLS.items():
        defaults[field] = ws[cell].value
    wb.close()
    return defaults


@app.post("/calculate")
def calculate(inputs: ModelInputs):
    """
    Write inputs to green cells only, trigger recalculation,
    read outputs. Excel model is not modified — a temp copy is used.
    """
    if not EXCEL_TEMPLATE.exists():
        raise HTTPException(404, "Excel template not found. Place model at backend/Spinnakeri_uürimudel_v3.xlsx")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        shutil.copy2(EXCEL_TEMPLATE, tmp_path)
        wb = load_workbook(tmp_path)
        ws = wb[SHEET]

        # Write ONLY green input cells
        input_values = inputs.model_dump()
        for field, cell_addr in INPUT_CELLS.items():
            if field in input_values:
                ws[cell_addr] = input_values[field]

        wb.save(tmp_path)
        wb.close()

        outputs = recalculate_excel(tmp_path)
        return {"inputs": input_values, "outputs": outputs}

    finally:
        os.unlink(tmp_path)


@app.get("/health")
def health():
    return {"status": "ok", "excel_template_found": EXCEL_TEMPLATE.exists()}
